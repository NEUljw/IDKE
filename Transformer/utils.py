import os
import torch
import pickle as pkl
from openpyxl import load_workbook
import jieba
import models.Transformer


def remove_blank(text):
    after = []
    for i in text:
        if i != ' ':
            after.append(i)
    return after


def build_vocab(file_path):
    # 建立词典
    vocab_dic = {}
    vocab = []
    vocab_num = 0

    wb = load_workbook(file_path)
    ws = wb.active
    rows = ws.rows
    for row in rows:
        line = [col.value for col in row]
        seq = line[0]
        word_list = jieba.lcut(seq)
        word_list = remove_blank(word_list)
        vocab += word_list
    vocab = list(set(vocab))
    for i in vocab:
        vocab_dic[i] = vocab_num
        vocab_num += 1

    vocab_dic['<UNK>'] = vocab_num
    vocab_num += 1
    vocab_dic['<PAD>'] = vocab_num
    vocab_num += 1

    return vocab_dic


def load_dataset(path, pad_size, vocab):
    # 读取数据
    contents = []

    wb = load_workbook(path)
    ws = wb.active
    rows = ws.rows
    for row in rows:
        line = [col.value for col in row]
        seq = line[0]
        label = line[1]
        word_list = jieba.lcut(seq)
        word_list = remove_blank(word_list)
        useful_seq_len = len(word_list)

        if len(word_list) < pad_size:
            word_list.extend(['<PAD>'] * (pad_size - len(word_list)))
        else:
            word_list = word_list[:pad_size]
            useful_seq_len = pad_size

        words_line = []
        # word to id
        for word in word_list:
            words_line.append(vocab.get(word, vocab.get('<UNK>')))

        contents.append((words_line, int(label), useful_seq_len))

    return contents


def build_dataset(config):
    # 有词典就加载词典，否则生成词典并储存
    if os.path.exists(config.vocab_path):
        vocab = pkl.load(open(config.vocab_path, 'rb'))
        print('load vocab done.')
    else:
        vocab = build_vocab(config.train_path)
        pkl.dump(vocab, open(config.vocab_path, 'wb'))
        print('save vocab done.')

    print("vocab size:", len(vocab))
    # 读取训练集，验证集和测试集
    train = load_dataset(config.train_path, config.pad_size, vocab=vocab)
    dev_and_test = load_dataset(config.dev_and_test_path, config.pad_size, vocab=vocab)
    dev = dev_and_test[:len(dev_and_test)//2]
    test = dev_and_test[len(dev_and_test)//2:]
    return vocab, train, dev, test


class DatasetIterater(object):
    def __init__(self, batches, batch_size):
        self.batch_size = batch_size
        self.batches = batches
        self.n_batches = len(batches) // batch_size
        self.residue = False  # 记录batch数量是否为整数
        if len(batches) % self.n_batches != 0:
            self.residue = True
        self.index = 0

    def _to_tensor(self, datas):
        x = torch.LongTensor([_[0] for _ in datas])
        y = torch.LongTensor([_[1] for _ in datas])

        # pad前的长度(超过pad_size的设为pad_size)
        seq_len = torch.LongTensor([_[2] for _ in datas])
        return (x, seq_len), y

    def __next__(self):
        if self.residue and self.index == self.n_batches:
            batches = self.batches[self.index * self.batch_size: len(self.batches)]
            self.index += 1
            batches = self._to_tensor(batches)
            return batches

        elif self.index >= self.n_batches:
            self.index = 0
            raise StopIteration
        else:
            batches = self.batches[self.index * self.batch_size: (self.index + 1) * self.batch_size]
            self.index += 1
            batches = self._to_tensor(batches)
            return batches

    def __iter__(self):
        return self

    def __len__(self):
        if self.residue:
            return self.n_batches + 1
        else:
            return self.n_batches


def build_iterator(dataset, config):
    iter = DatasetIterater(dataset, config.batch_size)
    return iter


if __name__ == "__main__":
    config = models.Transformer.Config()
    v, train, dev, test = build_dataset(config)
    print(v)
    print(train)
    print(dev)
    print(test)
