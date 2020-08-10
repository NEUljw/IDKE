import jieba
from openpyxl import load_workbook


def read_train_data(file_path, data_number):
    src_list, tgt_list = [], []
    wb = load_workbook(file_path)
    ws = wb.active
    rows = ws.rows
    n = 0
    for row in rows:
        n += 1
        if n <= data_number:
            line = [col.value for col in row]
            src_sen = line[0]
            tgt_sen = line[1]
            src_list.append(src_sen)
            tgt_list.append(tgt_sen)
    return src_list, tgt_list


def read_dev_data(file_path, data_number):
    src_list, endings_list, answer_list = [], [], []
    wb = load_workbook(file_path)
    ws = wb.active
    rows = ws.rows
    n = 0
    for row in rows:
        n += 1
        if n <= data_number:
            line = [col.value for col in row]
            src_sen = line[0]
            endings = line[1:5]
            answer = line[5]
            src_list.append(src_sen)
            endings_list += endings
            answer_list.append(answer)
    return src_list, endings_list, answer_list


def remove_blank(text):
    after = []
    for i in text:
        if i != ' ':
            after.append(i)
    return after


def build_vocab(sentences, vocab_type):
    vocab_list = []
    for sen in sentences:
        word_list = jieba.lcut(sen)
        word_list = remove_blank(word_list)
        vocab_list += word_list
    vocab_list = list(set(vocab_list))

    vocab_dict = dict()
    # Padding Should be Zero index
    vocab_dict['P'] = 0
    n = 1
    for word in vocab_list:
        vocab_dict[word] = n
        n += 1
    # source句子的词典
    if vocab_type == 'src':
        return vocab_dict
    # target句子的词典，需要加上S、E
    else:
        vocab_dict['S'] = n
        n += 1
        vocab_dict['E'] = n
        return vocab_dict


def padding_list(words, max_seq_len):
    if len(words) >= max_seq_len:
        return words[:max_seq_len]
    else:
        words += [0 for _ in range(max_seq_len - len(words))]
        return words


def coding_sentences(max_seq_len, train_data_number, dev_data_number):
    pre_sentences, ending = read_train_data(file_path='data/train.xlsx', data_number=train_data_number)
    pres, ends, answers = read_dev_data(file_path='data/dev.xlsx', data_number=dev_data_number)

    src_vocab = build_vocab(pre_sentences + pres, vocab_type='src')
    tgt_vocab = build_vocab(ending + ends, vocab_type='tgt')

    src_int = []
    for src_sen in pre_sentences:
        word_list = jieba.lcut(src_sen)
        word_list = remove_blank(word_list)
        src_sen_int = [src_vocab[i] for i in word_list]
        src_sen_int = padding_list(src_sen_int, max_seq_len=max_seq_len)
        src_int.append(src_sen_int)

    tgt_int, output_int = [], []
    for tgt_sen in ending:
        word_list = jieba.lcut(tgt_sen)
        word_list = remove_blank(word_list)
        tgt_sen_int = [tgt_vocab[i] for i in word_list]
        tgt_sen_int1 = [tgt_vocab['S']] + tgt_sen_int
        tgt_sen_int2 = tgt_sen_int + [tgt_vocab['E']]
        tgt_sen_int1 = padding_list(tgt_sen_int1, max_seq_len=max_seq_len)
        tgt_sen_int2 = padding_list(tgt_sen_int2, max_seq_len=max_seq_len)
        # 防止padding时把句子的结束标识符删除
        if tgt_sen_int2[-1] != 0 and tgt_sen_int2[-1] != tgt_vocab['E']:
            tgt_sen_int2[-1] = tgt_vocab['E']
        tgt_int.append(tgt_sen_int1)
        output_int.append(tgt_sen_int2)

    dev_src_int = []
    for dev_src_sen in pres:
        word_list = jieba.lcut(dev_src_sen)
        word_list = remove_blank(word_list)
        src_sen_int = [src_vocab[i] for i in word_list]
        src_sen_int = padding_list(src_sen_int, max_seq_len=max_seq_len)
        dev_src_int.append(src_sen_int)

    dev_endings_int, dev_output_int = [], []
    for dev_end in ends:
        word_list = jieba.lcut(dev_end)
        word_list = remove_blank(word_list)
        tgt_sen_int = [tgt_vocab[i] for i in word_list]
        tgt_sen_int1 = [tgt_vocab['S']] + tgt_sen_int
        tgt_sen_int2 = tgt_sen_int + [tgt_vocab['E']]
        tgt_sen_int1 = padding_list(tgt_sen_int1, max_seq_len=max_seq_len)
        tgt_sen_int2 = padding_list(tgt_sen_int2, max_seq_len=max_seq_len)
        # 防止padding时把句子的结束标识符删除
        if tgt_sen_int2[-1] != 0 and tgt_sen_int2[-1] != tgt_vocab['E']:
            tgt_sen_int2[-1] = tgt_vocab['E']
        dev_endings_int.append(tgt_sen_int2)
        dev_output_int.append(tgt_sen_int1)
    return src_vocab, tgt_vocab, src_int, tgt_int, output_int, dev_src_int, dev_endings_int, dev_output_int, answers
